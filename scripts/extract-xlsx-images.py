#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import openpyxl


def slugify(text: str) -> str:
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "wine"


def pick_extension(img) -> str:
    path = getattr(img, "path", "") or ""
    suffix = Path(path).suffix.lower()
    if suffix in {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}:
        return suffix
    return ".png"


def image_row(img) -> int | None:
    anchor = getattr(img, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    # openpyxl uses zero-based row index in anchors
    return int(marker.row) + 1


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract embedded images from XLSX and map them to worksheet row numbers.")
    parser.add_argument("xlsx_path", help="Path to source .xlsx file")
    parser.add_argument("--sheet", default="vein", help="Worksheet name (default: vein)")
    parser.add_argument("--out-dir", default="data/extracted-images", help="Directory to save extracted images")
    parser.add_argument("--map-json", default="data/image-row-map.json", help="Output JSON mapping path")
    parser.add_argument("--name-col", default="Nimi", help="Header used for naming files (default: Nimi)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx_path)
    out_dir = Path(args.out_dir)
    map_json = Path(args.map_json)

    if not xlsx_path.exists():
      raise SystemExit(f"XLSX file not found: {xlsx_path}")

    out_dir.mkdir(parents=True, exist_ok=True)
    map_json.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if args.sheet not in wb.sheetnames:
      raise SystemExit(f"Sheet not found: {args.sheet}. Available: {wb.sheetnames}")

    ws = wb[args.sheet]

    header_row = 2
    headers = {}
    for col in range(1, ws.max_column + 1):
      value = ws.cell(row=header_row, column=col).value
      if value is not None:
        headers[str(value).strip()] = col

    name_col = headers.get(args.name_col)

    mapping = {
      "xlsx": str(xlsx_path),
      "sheet": args.sheet,
      "items": [],
    }

    images = getattr(ws, "_images", [])
    for index, img in enumerate(images, start=1):
      row = image_row(img)
      if row is None:
        continue

      wine_name = ""
      if name_col:
        raw_name = ws.cell(row=row, column=name_col).value
        wine_name = "" if raw_name is None else str(raw_name)

      ext = pick_extension(img)
      file_name = f"r{row:04d}-{slugify(wine_name)}-{index:03d}{ext}"
      file_path = out_dir / file_name

      data = img._data() if callable(getattr(img, "_data", None)) else None
      if not data:
        continue

      file_path.write_bytes(data)
      mapping["items"].append(
        {
          "row": row,
          "wineName": wine_name,
          "file": str(file_path),
        }
      )

    map_json.write_text(json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Extracted images: {len(mapping['items'])}")
    print(f"Images folder: {out_dir}")
    print(f"Mapping JSON: {map_json}")


if __name__ == "__main__":
    main()
